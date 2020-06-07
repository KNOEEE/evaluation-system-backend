# -*- coding: utf-8 -*-
from django.http import HttpResponse
from django.shortcuts import render
from Model.models import student, comment, course, join, favor
import json
import tkinter as tk  # 打开文件对话框
from tkinter import filedialog
import os
import tkinter.messagebox
import openpyxl
from openpyxl import load_workbook
from django.contrib import messages


def hello(request):
    return render(request, 'hello1.html')  # 'base.html'


def home(request):
    request.encoding = 'utf-8'
    aid = request.GET['aid']
    pwd = request.GET['pwd']
    # req = json.loads(request.body)
    # aid = req['aid']
    # pwd = req['pwd']
    if (aid == 'admin') and (pwd == 'admin'):
        return render(request, 'hello.html')
    else:
        return HttpResponse("账号或密码错误")


def showTable(request):
    request.encoding = 'utf-8'
    tableWant = request.GET['table']
    dict1 = {'student': student,
             'course': course,
             'join': join,
             'favor': favor,
             'comm': comment}
    # list1 = dict1[tableWant].objects.all().order_by('sid')
    dataInTable = []
    if tableWant == 'student':
        list1 = student.objects.all().order_by('sid')
        for var in list1:
            dataLine = {"sid": var.sid, "pwd": var.pwd}
            dataInTable.append(dataLine)
        context = {"table": tableWant,
                   "datalist": dataInTable}
        return render(request, 'show/showStu.html', context)
    elif tableWant == 'course':
        list1 = course.objects.all().order_by('cid')
        for var in list1:
            dataLine = {'cid': var.cid, 'rate': var.rate,
                        'nums': var.nums, 'name': var.name,
                        'teacher': var.teacher}
            dataInTable.append(dataLine)
        context = {"table": tableWant,
                   "datalist": dataInTable}
        return render(request, 'show/showCourse.html', context)
    elif tableWant == 'join':
        list1 = join.objects.all().order_by('cid')
        for var in list1:
            dataLine = {'cid': var.cid, 'rate': var.rate,
                        'sid': var.sid, 'cname': var.cname,
                        'teacher': var.teacher}
            dataInTable.append(dataLine)
        context = {"table": tableWant,
                   "datalist": dataInTable}
        return render(request, 'show/showJoin.html', context)
    elif tableWant == 'favor':
        list1 = favor.objects.all().order_by('sid')
        for var in list1:
            dataLine = {'cid': var.cid,
                        'sid': var.sid, 'cname': var.cname,
                        'teacher': var.teacher}
            dataInTable.append(dataLine)
        context = {"table": tableWant,
                   "datalist": dataInTable}
        return render(request, 'show/showFav.html', context)
    elif tableWant == 'review':
        list1 = comment.objects.all().order_by('cid')
        for var in list1:
            dataLine = {'cid': var.cid, 'rate': var.rate,
                        'sid': var.sid, 'text': var.text,
                        'index1': var.index1, 'index2': var.index2,
                        'index3': var.index3, 'index4': var.index4,
                        'index5': var.index5}
            dataInTable.append(dataLine)
        context = {"table": tableWant,
                   "datalist": dataInTable}
        return render(request, 'show/showComm.html', context)
    else:
        return HttpResponse("抱歉，您要找的页面不存在")


def addStBatch(request):
    if request.method == 'POST':
        File = request.FILES.get("myfile", None)
        if File is None:
            return HttpResponse("没有要上传的文件")
        else:
            tempFile = './myServer/tempFile/temp.xlsx'
            if os.path.exists(tempFile):
                os.remove(tempFile)
            with open(tempFile, 'wb+') as f:  # % File.name
                for chunk in File.chunks():
                    f.write(chunk)

            try:
                wb = load_workbook(tempFile, True)
            except:
                return HttpResponse("文件错误")
            ws = wb.active
            maxRow = ws.max_row
            try:
                for row in ws.iter_rows(min_row=2, min_col=1,
                                        max_row=maxRow, max_col=2):
                    sid = row[0].value
                    pwd = row[1].value
                    opr = student(sid=sid, pwd=pwd)
                    opr.save()
                # del wb, ws
                # gc.collect()
                wb.close()
            except Exception as e:
                wb.close()
                return HttpResponse("数据异常")
            return HttpResponse("导入成功!")
    else:
        return render(request, "show/showStu.html")


def addLine(request):
    request.encoding = 'utf-8'

    if request.GET['addType'] == '添加学生':
        sid = request.GET['sid']
        pwd = request.GET['pwd']
        if (sid == '') or (pwd == ''):
            return HttpResponse("请输入信息")
        try:
            opr = student(sid=sid, pwd=pwd)
            opr.save()
        except Exception as e:
            return HttpResponse("数据异常")
    elif request.GET['addType'] == '添加课程':
        cid = request.GET['cid']
        name = request.GET['name']
        teacher = request.GET['teacher']
        if (name == '') or (cid == '') or (teacher == ''):
            return HttpResponse("请输入信息")
        try:
            opr = course(cid=cid, name=name, rate=0,
                         teacher=teacher, nums=0)
            opr.save()
        except Exception as e:
            return HttpResponse("数据异常")
    elif request.GET['addType'] == '添加':
        sid = request.GET['sid']
        cid = request.GET['cid']
        if (sid == '') or (cid == ''):
            return HttpResponse("请输入信息")
        try:
            courseAttend = course.objects.get(cid=cid)
            studentAttend = student.objects.get(sid=sid)  # 看stu表里有没有要添加的学生
        except Exception as e:
            return HttpResponse("课程或学生不存在")
        name = courseAttend.name
        teacher = courseAttend.teacher
        try:
            opr = join(sid=sid, cid=cid, cname=name,
                       teacher=teacher, rate=0)
            opr.save()
        except Exception as e:
            return HttpResponse("数据异常")
    else:
        return HttpResponse("抱歉，您要找的页面不存在")
    return HttpResponse("上传成功")


def delStLine(request):
    request.encoding = 'utf-8'
    sid = request.GET['sid']
    studentDel = student.objects.filter(sid=sid)
    if len(studentDel) == 0:
        return HttpResponse("该学生不存在")
    list1 = join.objects.filter(sid=sid)
    for var in list1:
        if var.rate > 0:
            courseNew = course.objects.get(cid=var.cid)
            rateNew = (courseNew.nums * courseNew.rate - var.rate) / (courseNew.nums - 1)
            courseNew.rate = rateNew
            courseNew.nums = courseNew.nums - 1
            courseNew.save()
    studentDel.delete()
    comment.objects.filter(sid=sid).delete()
    join.objects.filter(sid=sid).delete()
    favor.objects.filter(sid=sid).delete()
    return HttpResponse("删除成功")


def delCsLine(request):
    request.encoding = 'utf-8'
    cid = request.GET['cid']
    courseDel = course.objects.filter(cid=cid)
    if len(courseDel) == 0:
        return HttpResponse("课程不存在")
    courseDel.delete()
    favor.objects.filter(cid=cid).delete()
    join.objects.filter(cid=cid).delete()
    comment.objects.filter(cid=cid).delete()
    return HttpResponse("删除成功")


def delJnLine(request):
    request.encoding = 'utf-8'
    cid = request.GET['cid']
    sid = request.GET['sid']  # 学生离开该课程后，评价仍然保留
    joinDel = join.objects.filter(sid=sid, cid=cid)
    if len(joinDel) == 0:
        return HttpResponse("课程或学生不存在")
    joinDel.delete()
    return HttpResponse("删除成功")


def delCmLine(request):
    request.encoding = 'utf-8'
    cid = request.GET['cid']
    sid = request.GET['sid']
    rate = request.GET['rate']
    rate = float(rate)  # 在所有删除函数都要考虑str转float，及删除后除数变为0
    commDel = comment.objects.filter(sid=sid, cid=cid)
    # joinDel = join.objects.filter(sid=sid, cid=cid)
    if len(commDel) == 0:
        return HttpResponse("评价不存在")

    joinNew = join.objects.get(sid=sid, cid=cid)
    joinNew.rate = 0
    joinNew.save()

    courseNew = course.objects.get(cid=cid)
    if courseNew.nums == 1:
        rateNew = 0
    else:
        rateNew = (courseNew.nums * courseNew.rate - rate) / (courseNew.nums - 1)
    courseNew.rate = rateNew
    courseNew.nums = courseNew.nums - 1
    courseNew.save()

    commDel.delete()
    return HttpResponse("删除成功")


def addCsBatch(request):
    if request.method == 'POST':
        File = request.FILES.get("myfile", None)
        if File is None:
            return HttpResponse("没有要上传的文件")
        else:
            tempFile = './myServer/tempFile/temp.xlsx'
            if os.path.exists(tempFile):
                os.remove(tempFile)
            with open(tempFile, 'wb+') as f:  # % File.name
                for chunk in File.chunks():
                    f.write(chunk)

            try:
                wb = load_workbook(tempFile, True)
            except:
                return HttpResponse("文件错误")
            ws = wb.active
            maxRow = ws.max_row
            try:
                for row in ws.iter_rows(min_row=2, min_col=1,
                                        max_row=maxRow, max_col=3):
                    cid = row[0].value
                    name = row[1].value
                    if row[2].value is None:
                        teacher = '-'
                    else:
                        teacher = row[2].value
                    opr = course(cid=cid, name=name, teacher=teacher,
                                 rate=0, nums=0)
                    opr.save()
                wb.close()
            except Exception as e:
                wb.close()
                return HttpResponse("数据异常")
            return HttpResponse("导入成功!")
    else:
        return render(request, "show/showStu.html")


def addJnBatch(request):
    if request.method == 'POST':
        File = request.FILES.get("myfile", None)
        if File is None:
            return HttpResponse("没有要上传的文件")
        else:
            tempFile = './myServer/tempFile/temp.xlsx'
            if os.path.exists(tempFile):
                os.remove(tempFile)
            with open(tempFile, 'wb+') as f:  # % File.name
                for chunk in File.chunks():
                    f.write(chunk)

            try:
                wb = load_workbook(tempFile, True)  # 处理上传的文件不是excel
            except:
                return HttpResponse("文件错误")
            ws = wb.active
            try:
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    sid = row[0].value
                    studentAttendList = student.objects.filter(sid=sid)
                    if len(studentAttendList) == 0:
                        wb.close()
                        return HttpResponse("数据中包含的学生不存在%d" % ws.max_row)
                # del row[0]
                    for cell in row[1:]:  # 从每行的第二列开始是参与的课程号
                        if cell.value is None:
                            break  # 一行结束
                        courseAttendList = course.objects.filter(cid=cell.value)
                        if len(courseAttendList) == 0:
                            wb.close()
                            return HttpResponse("该课程不存在: " + cell.value)  # "数据中包含的课程不存在"

                        joinToAdd = join.objects.filter(sid=sid, cid=cell.value)
                        if len(joinToAdd) == 1:
                            continue  # 重复的参与关系
                        courseAttend = courseAttendList[0]
                        cname = courseAttend.name
                        teacher = courseAttend.teacher
                        opr = join(sid=sid, cid=cell.value, cname=cname,
                                   teacher=teacher, rate=0)
                        opr.save()
                wb.close()
            except Exception as e:
                wb.close()
                return HttpResponse("数据异常")
            return HttpResponse("导入成功!")
    else:
        return render(request, "show/showJoin.html")
